import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib


#setando a aparência padrão do sistema
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.appearance()
        self.todo_sistema()
       

    def layout_config(self):
        self.title("Sistema de Gestão de Clientes")
        self.geometry("700x500")


    def appearance(self):
        self.lb_apm = ctk.CTkLabel(self, text="Tema", bg_color="transparent", text_color=['#000','#fff']).place(x=50,y=430)
        self.opt_apm = ctk.CTkOptionMenu(self, values=['Light','Dark','System'], command=self.change_apm).place(x=50,y=460)

    def todo_sistema(self):
        frame = ctk.CTkFrame(self, width=700, height=50, corner_radius=0, bg_color="teal", fg_color="teal")
        frame.place(x=0,y=10)
        title = ctk.CTkLabel(frame, text="Sistema de Gestão de Clientes", font=("Century Gothic bold",24), text_color="#fff").place(x=190,y=10)

        span = ctk.CTkLabel(self, text="Por favor, preencha os campos do formulário!", font=("Century Gothic bold",16), text_color=["#000","#fff"]).place(x=50,y=70)
        
        ficheiro = pathlib.Path('clientes.xlsx')

        if ficheiro.exists():
            pass
        else:
           ficheiro=Workbook()
           folha = ficheiro.active
           folha['A1']='Nome Completo'
           folha['B1']= 'Contato'
           folha['C1']= 'Idade'
           folha['D1']= 'Gênero'
           folha['E1']= 'Endereço'
           folha['F1']= 'Observações'

           ficheiro.save('clientes.xlsx')


        def submit():
        
            #pegando os dados dos entrys
            nome = nome_value.get()
            contato = contato_value.get()
            idade = idade_value.get()
            genero = genero_combobox.get()
            endereco = endereco_value.get()
            obs = obs_entry.get(0.0,END)

            if (nome =='' or contato=='' or idade=='' or endereco==''):
                messagebox.showerror('Sistema', 'ERRO!\n Por favor, preencha todos os campos.')
            else:    
  #lançar dados no sistema
                ficheiro = openpyxl.load_workbook('clientes.xlsx')
                folha = ficheiro.active
                folha.cell(column=1,row=folha.max_row+1,value=nome)
                folha.cell(column=2,row=folha.max_row,value=contato)
                folha.cell(column=3,row=folha.max_row,value=idade)
                folha.cell(column=4,row=folha.max_row,value=genero)
                folha.cell(column=5,row=folha.max_row,value=endereco)
                folha.cell(column=6,row=folha.max_row,value=obs)

                ficheiro.save(r'clientes.xlsx')
                messagebox.showinfo("Sistema", "Dados Salvos com sucesso!")




        def clear():
            nome_value.set("")
            contato_value.set("")
            idade_value.set("")
            genero_combobox.set("")
            obs_entry.delete(0.0,END)


        #text variaveis
        nome_value = StringVar()
        contato_value = StringVar()
        idade_value = StringVar()
        endereco_value = StringVar()

        #Entrys
        nome_entry = ctk.CTkEntry(self, width=350, textvariable=nome_value, font=("Century Gothic bold", 16),fg_color="transparent")
        contato_entry = ctk.CTkEntry(self, width=200, textvariable=contato_value, font=("Century Gothic bold", 16),fg_color="transparent")
        idade_entry = ctk.CTkEntry(self, width=150, textvariable=idade_value, font=("Century Gothic bold", 16),fg_color="transparent")
        endereco_entry = ctk.CTkEntry(self, width=200,textvariable=endereco_value, font=("Century Gothic bold", 16),fg_color="transparent")

        #Combobox
        genero_combobox = ctk.CTkComboBox(self,values=['Masculino', 'Feminino'],font=('Century Gothic bold', 14),width=150)
        genero_combobox.set('Masculino')

        #entrada de observações
        obs_entry = ctk.CTkTextbox(self,width=500, height=150, font=("arial",18),border_color="#aaa",border_width=2, fg_color="transparent")

         #Labels
        lb_nome = ctk.CTkLabel(self, text="Nome Completo", font=("Century Gothic bold",14), text_color=["#000","#fff"])
        lb_contato = ctk.CTkLabel(self, text="Contato", font=("Century Gothic bold",14), text_color=["#000","#fff"])
        lb_idade = ctk.CTkLabel(self, text="Idade", font=("Century Gothic bold",14), text_color=["#000","#fff"])
        lb_genero = ctk.CTkLabel(self, text="Gênero", font=("Century Gothic bold",14), text_color=["#000","#fff"])
        lb_endereco = ctk.CTkLabel(self, text="Endereço", font=("Century Gothic bold",14), text_color=["#000","#fff"])
        lb_obs = ctk.CTkLabel(self, text="Observações", font=("Century Gothic bold",14), text_color=["#000","#fff"])

        btn_submit = ctk.CTkButton(self, text="Salvar dados".upper(),command=submit, fg_color="#151",hover_color="#131").place(x=300,y=420)
        btn_submit = ctk.CTkButton(self, text="Limpar campos".upper(),command=clear, fg_color="#555",hover_color="#333").place(x=500,y=420)


        #posicionando os elementos na janela
        lb_nome.place(x=50,y=120)
        nome_entry.place(x=50,y=150)

        lb_contato.place(x=450,y=120)
        contato_entry.place(x=450, y=150)

        lb_idade.place(x=300,y=190)
        idade_entry.place(x=300,y=220)

        lb_genero.place(x=500,y=190)
        genero_combobox.place(x=500, y=220)

        lb_endereco.place(x=50, y=190)
        endereco_entry.place(x=50, y=220)

        lb_obs.place(x=50, y=260)
        obs_entry.place(x=150,y=260)




    def change_apm(self, nova_aparencia):
        ctk.set_appearance_mode(nova_aparencia)

if __name__=="__main__":
    app = App()
    app.mainloop()
